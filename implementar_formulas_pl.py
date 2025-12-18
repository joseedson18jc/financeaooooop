#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Implementar fórmulas SUMIFS e cálculos automáticos no P&L
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from dateutil.relativedelta import relativedelta

print("=" * 80)
print("IMPLEMENTANDO FÓRMULAS AUTOMÁTICAS NO P&L")
print("=" * 80)

# ============================================================================
# 1. CARREGAR WORKBOOK
# ============================================================================
print("\n1. Carregando workbook...")

wb = load_workbook('/home/ubuntu/Business_Plan_Umatch_Automatizado_v2.xlsx')
ws_pl = wb['P&L']
ws_extrato = wb['Extrato_Importado']

print(f"   ✓ Workbook carregado")

# ============================================================================
# 2. GERAR MESES PARA REFERÊNCIA
# ============================================================================
print("\n2. Gerando referências de meses...")

data_inicio = datetime(2024, 5, 1)
meses_ref = []
for i in range(18):
    mes = data_inicio + relativedelta(months=i)
    meses_ref.append({
        'col_idx': i + 3,
        'col_letter': get_column_letter(i + 3),
        'mes_str': mes.strftime("%m/%Y"),
        'mes_num': mes.month,
        'ano': mes.year,
        'periodo': mes.strftime("%Y-%m")
    })

print(f"   ✓ {len(meses_ref)} meses configurados")

# ============================================================================
# 3. IMPLEMENTAR FÓRMULAS DE IMPORTAÇÃO (SUMIFS)
# ============================================================================
print("\n3. Implementando fórmulas SUMIFS para importação...")

# Mapeamento de linhas para importação
# Formato: {linha: {'centro_custo': '', 'fornecedor': '', 'tipo': ''}}
mapeamento_importacao = {
    # RECEITAS
    29: {  # Google
        'centro_custo': 'Google Play Net Revenue',
        'fornecedor': None,  # Todos
        'tipo': 'receita'
    },
    30: {  # Google - Brazil
        'centro_custo': 'Google Play Net Revenue',
        'fornecedor': None,
        'tipo': 'receita',
        'filtro_adicional': 'Brazil'
    },
    32: {  # Google - USA
        'centro_custo': 'Google Play Net Revenue',
        'fornecedor': None,
        'tipo': 'receita',
        'filtro_adicional': 'USA'
    },
    37: {  # Apple
        'centro_custo': 'App Store Net Revenue',
        'fornecedor': None,
        'tipo': 'receita'
    },
    38: {  # Apple - Brazil
        'centro_custo': 'App Store Net Revenue',
        'fornecedor': None,
        'tipo': 'receita',
        'filtro_adicional': 'Brazil'
    },
    40: {  # Apple - USA
        'centro_custo': 'App Store Net Revenue',
        'fornecedor': None,
        'tipo': 'receita',
        'filtro_adicional': 'USA'
    },
    42: {  # Invest Income
        'centro_custo': 'Rendimentos de Aplicações',
        'fornecedor': None,
        'tipo': 'receita'
    },
    
    # COGS
    47: {  # AWS
        'centro_custo': 'Web Services Expenses',
        'fornecedor': 'AWS',
        'tipo': 'custo'
    },
    48: {  # Cloudflare
        'centro_custo': 'Web Services Expenses',
        'fornecedor': 'Cloudflare',
        'tipo': 'custo'
    },
    49: {  # Heroku
        'centro_custo': 'Web Services Expenses',
        'fornecedor': 'Heroku',
        'tipo': 'custo'
    },
    50: {  # IAPHUB
        'centro_custo': 'Web Services Expenses',
        'fornecedor': 'IAPHUB',
        'tipo': 'custo'
    },
    51: {  # MailGun
        'centro_custo': 'Web Services Expenses',
        'fornecedor': 'MailGun',
        'tipo': 'custo'
    },
    52: {  # AWS SES
        'centro_custo': 'Web Services Expenses',
        'fornecedor': 'AWS SES',
        'tipo': 'custo'
    },
    
    # SG&A
    60: {  # Marketing
        'centro_custo': 'Marketing & Growth Expenses',
        'fornecedor': None,
        'tipo': 'custo'
    },
    68: {  # Wages
        'centro_custo': 'Wages Expenses',
        'fornecedor': None,
        'tipo': 'custo'
    },
    69: {  # Tech Support & Services (total)
        'centro_custo': 'Tech Support & Services',
        'fornecedor': None,
        'tipo': 'custo'
    },
}

# Implementar fórmulas SUMIFS
contador_formulas = 0
for linha, config in mapeamento_importacao.items():
    for mes_info in meses_ref:
        col_letter = mes_info['col_letter']
        cell_ref = f'{col_letter}{linha}'
        
        # Construir fórmula SUMIFS
        # =SUMIFS(Extrato_Importado!$E:$E, Extrato_Importado!$H:$H, "2024-09", Extrato_Importado!$B:$B, "Centro de Custo")
        
        formula_parts = [
            f'SUMIFS(Extrato_Importado!$E:$E',  # Soma coluna Valor
            f'Extrato_Importado!$H:$H',  # Critério: Mês
            f'"{mes_info["periodo"]}"',  # Valor do mês
            f'Extrato_Importado!$B:$B',  # Critério: Centro de Custo
            f'"{config["centro_custo"]}"'  # Valor do centro de custo
        ]
        
        # Adicionar filtro de fornecedor se especificado
        if config.get('fornecedor'):
            formula_parts.extend([
                f'Extrato_Importado!$C:$C',  # Critério: Fornecedor/Cliente
                f'"*{config["fornecedor"]}*"'  # Valor do fornecedor (com wildcards)
            ])
        
        formula = ','.join(formula_parts) + ')'
        
        ws_pl[cell_ref] = f'={formula}'
        contador_formulas += 1

print(f"   ✓ {contador_formulas} fórmulas SUMIFS implementadas")

# ============================================================================
# 4. IMPLEMENTAR FÓRMULAS DE CÁLCULO
# ============================================================================
print("\n4. Implementando fórmulas de cálculo...")

contador_calc = 0

# Para cada mês
for mes_info in meses_ref:
    col = mes_info['col_letter']
    
    # Revenue (linha 24) = soma de todas as receitas
    ws_pl[f'{col}24'] = f'={col}25+{col}42'
    contador_calc += 1
    
    # Revenue no Tax (linha 25) = Google + Apple
    ws_pl[f'{col}25'] = f'={col}29+{col}37'
    contador_calc += 1
    
    # Google (IAPHUB) x 0.85 (linha 27)
    ws_pl[f'{col}27'] = f'={col}26*0.85'
    contador_calc += 1
    
    # Δ(%) Google (linha 28)
    if mes_info['col_idx'] > 3:  # A partir do segundo mês
        col_anterior = get_column_letter(mes_info['col_idx'] - 1)
        ws_pl[f'{col}28'] = f'=IFERROR(({col}27-{col_anterior}27)/{col_anterior}27,0)'
    else:
        ws_pl[f'{col}28'] = 0
    contador_calc += 1
    
    # Brazil Margin (%) Google (linha 31)
    ws_pl[f'{col}31'] = f'=IFERROR({col}30/{col}29,0)'
    contador_calc += 1
    
    # USA Margin (%) Google (linha 33)
    ws_pl[f'{col}33'] = f'=IFERROR({col}32/{col}29,0)'
    contador_calc += 1
    
    # Apple (IAPHUB) x 0.85 (linha 35)
    ws_pl[f'{col}35'] = f'={col}34*0.85'
    contador_calc += 1
    
    # Δ(%) Apple (linha 36)
    if mes_info['col_idx'] > 3:
        col_anterior = get_column_letter(mes_info['col_idx'] - 1)
        ws_pl[f'{col}36'] = f'=IFERROR(({col}35-{col_anterior}35)/{col_anterior}35,0)'
    else:
        ws_pl[f'{col}36'] = 0
    contador_calc += 1
    
    # Brazil Margin (%) Apple (linha 39)
    ws_pl[f'{col}39'] = f'=IFERROR({col}38/{col}37,0)'
    contador_calc += 1
    
    # USA Margin (%) Apple (linha 41)
    ws_pl[f'{col}41'] = f'=IFERROR({col}40/{col}37,0)'
    contador_calc += 1
    
    # Payment Processing Expenses (linha 45) = 17.65% da Revenue no Tax
    ws_pl[f'{col}45'] = f'={col}25*0.1765'
    contador_calc += 1
    
    # COGS (linha 46) = soma de AWS a AWS SES
    ws_pl[f'{col}46'] = f'=SUM({col}47:{col}52)'
    contador_calc += 1
    
    # Costs of Revenue (linha 44) = Payment Processing + COGS
    ws_pl[f'{col}44'] = f'={col}45+{col}46'
    contador_calc += 1
    
    # Gross Profit (linha 54) = Revenue - Costs of Revenue
    ws_pl[f'{col}54'] = f'={col}24-{col}44'
    contador_calc += 1
    
    # Gross Margin (linha 55) = Gross Profit / Revenue
    ws_pl[f'{col}55'] = f'=IFERROR({col}54/{col}24,0)'
    contador_calc += 1
    
    # Marketing/Revenue no Tax (linha 61)
    ws_pl[f'{col}61'] = f'=IFERROR({col}60/{col}25,0)'
    contador_calc += 1
    
    # Tech Support & Services (linha 69) = soma dos itens
    # Por enquanto, usar apenas a importação direta
    # ws_pl[f'{col}69'] já tem SUMIFS
    
    # SG&A (linha 59) = Marketing + Wages + Tech Support
    ws_pl[f'{col}59'] = f'={col}60+{col}68+{col}69'
    contador_calc += 1
    
    # Operating Expenses (linha 57) = R&D + SG&A
    ws_pl[f'{col}57'] = f'={col}58+{col}59'
    contador_calc += 1

print(f"   ✓ {contador_calc} fórmulas de cálculo implementadas")

# ============================================================================
# 5. ADICIONAR LINHAS DE EBITDA E LUCRO LÍQUIDO
# ============================================================================
print("\n5. Adicionando linhas de EBITDA e resultado...")

# Encontrar última linha usada
ultima_linha = 72

# EBITDA
ws_pl[f'A{ultima_linha+1}'] = ""
ws_pl[f'A{ultima_linha+2}'] = "EBITDA"
ws_pl[f'B{ultima_linha+2}'] = "BRL"

# Operating Income (EBIT)
ws_pl[f'A{ultima_linha+3}'] = "Operating Income (EBIT)"
ws_pl[f'B{ultima_linha+3}'] = "BRL"

# EBITDA Margin
ws_pl[f'A{ultima_linha+4}'] = "     EBITDA Margin"
ws_pl[f'B{ultima_linha+4}'] = "%"

# Aplicar estilos
from openpyxl.styles import Font, PatternFill
profit_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
category_font = Font(bold=True, size=10)

ws_pl[f'A{ultima_linha+2}'].font = category_font
ws_pl[f'A{ultima_linha+2}'].fill = profit_fill
ws_pl[f'A{ultima_linha+3}'].font = category_font
ws_pl[f'A{ultima_linha+3}'].fill = profit_fill

# Fórmulas
for mes_info in meses_ref:
    col = mes_info['col_letter']
    
    # EBITDA (linha 74) = Gross Profit - Operating Expenses
    ws_pl[f'{col}{ultima_linha+2}'] = f'={col}54-{col}57'
    ws_pl[f'{col}{ultima_linha+2}'].number_format = 'R$ #,##0.00'
    
    # Operating Income (linha 75) = EBITDA (simplificado, sem D&A)
    ws_pl[f'{col}{ultima_linha+3}'] = f'={col}{ultima_linha+2}'
    ws_pl[f'{col}{ultima_linha+3}'].number_format = 'R$ #,##0.00'
    
    # EBITDA Margin (linha 76)
    ws_pl[f'{col}{ultima_linha+4}'] = f'=IFERROR({col}{ultima_linha+2}/{col}24,0)'
    ws_pl[f'{col}{ultima_linha+4}'].number_format = '0.00%'

print(f"   ✓ Linhas de EBITDA e resultado adicionadas")

# ============================================================================
# 6. SALVAR WORKBOOK
# ============================================================================
print("\n6. Salvando workbook...")

output_path = '/home/ubuntu/Business_Plan_Umatch_Automatizado_v3.xlsx'
wb.save(output_path)

print(f"\n✓ Workbook salvo em: {output_path}")
print("\n" + "=" * 80)
print("FASE 3 CONCLUÍDA - Fórmulas automáticas implementadas")
print("=" * 80)
print(f"\nTotal de fórmulas criadas: {contador_formulas + contador_calc + (18 * 3)}")
print("\nPróximas etapas:")
print("  - Criar aba DRE consolidado")
print("  - Criar aba Dashboard com gráficos")
print("  - Criar aba Glossário")
print("  - Criar aba Checklist")
