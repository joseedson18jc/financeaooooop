#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Adicionar aba P&L com fórmulas automáticas ao Business Plan
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime
from dateutil.relativedelta import relativedelta

print("=" * 80)
print("ADICIONANDO ABA P&L COM FÓRMULAS AUTOMÁTICAS")
print("=" * 80)

# ============================================================================
# 1. CARREGAR WORKBOOK EXISTENTE
# ============================================================================
print("\n1. Carregando workbook existente...")

wb = load_workbook('/home/ubuntu/Business_Plan_Umatch_Automatizado_v1.xlsx')
print(f"   ✓ Abas existentes: {wb.sheetnames}")

# Carregar P&L original para estrutura
pl_original = pd.read_csv('/home/ubuntu/upload/00_Business_Plan_Umatch.xlsx-P&L.csv')
print(f"   ✓ P&L original: {pl_original.shape}")

# ============================================================================
# 2. CRIAR ABA P&L
# ============================================================================
print("\n2. Criando aba P&L...")

ws_pl = wb.create_sheet("P&L", 0)  # Inserir como primeira aba

# Estilos
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
category_font = Font(bold=True, size=10)
subcategory_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
revenue_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
cost_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
editable_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================================
# 3. CABEÇALHO E ESTRUTURA DE MESES
# ============================================================================
print("\n3. Criando cabeçalho e estrutura de meses...")

# Título
ws_pl['A1'] = "BUSINESS PLAN UMATCH - P&L (REGIME DE COMPETÊNCIA)"
ws_pl['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_pl['A1'].fill = header_fill
ws_pl.merge_cells('A1:P1')

# Cabeçalhos de colunas
ws_pl['A3'] = "Categoria"
ws_pl['B3'] = "Unidade"

# Meses (últimos 12 meses + próximos 6 meses)
data_inicio = datetime(2024, 5, 1)
meses = []
for i in range(18):
    mes = data_inicio + relativedelta(months=i)
    meses.append(mes.strftime("%m/%Y"))

for col_idx, mes in enumerate(meses, start=3):
    col_letter = get_column_letter(col_idx)
    ws_pl[f'{col_letter}3'] = mes
    ws_pl[f'{col_letter}3'].font = header_font
    ws_pl[f'{col_letter}3'].fill = header_fill
    ws_pl[f'{col_letter}3'].alignment = Alignment(horizontal='center', vertical='center')
    ws_pl[f'{col_letter}3'].border = thin_border

ws_pl['A3'].font = header_font
ws_pl['A3'].fill = header_fill
ws_pl['A3'].alignment = Alignment(horizontal='center', vertical='center')
ws_pl['B3'].font = header_font
ws_pl['B3'].fill = header_fill
ws_pl['B3'].alignment = Alignment(horizontal='center', vertical='center')

# ============================================================================
# 4. ESTRUTURA DO P&L
# ============================================================================
print("\n4. Criando estrutura do P&L...")

# Estrutura de linhas do P&L
estrutura_pl = [
    # NAU (Net Active Users)
    {"linha": 5, "categoria": "NAU", "unidade": "Users", "tipo": "kpi", "nivel": 0},
    {"linha": 6, "categoria": "    Brazil", "unidade": "Users", "tipo": "kpi", "nivel": 1},
    {"linha": 7, "categoria": "        EaM", "unidade": "Users", "tipo": "kpi", "nivel": 2},
    {"linha": 8, "categoria": "        Affiliates & Events", "unidade": "Users", "tipo": "kpi", "nivel": 2},
    {"linha": 9, "categoria": "        Ads", "unidade": "Users", "tipo": "kpi", "nivel": 2},
    {"linha": 10, "categoria": "        Organic", "unidade": "Users", "tipo": "kpi", "nivel": 2},
    {"linha": 11, "categoria": "    USA", "unidade": "Users", "tipo": "kpi", "nivel": 1},
    {"linha": 12, "categoria": "        EaM", "unidade": "Users", "tipo": "kpi", "nivel": 2},
    {"linha": 13, "categoria": "        Ads", "unidade": "Users", "tipo": "kpi", "nivel": 2},
    {"linha": 14, "categoria": "        Organic", "unidade": "Users", "tipo": "kpi", "nivel": 2},
    {"linha": 15, "categoria": "", "unidade": "", "tipo": "blank", "nivel": 0},
    
    # CPA (Cost Per Acquisition)
    {"linha": 16, "categoria": "CPA", "unidade": "BRL", "tipo": "kpi", "nivel": 0},
    {"linha": 17, "categoria": "    Brazil", "unidade": "BRL", "tipo": "kpi", "nivel": 1},
    {"linha": 18, "categoria": "        EaM", "unidade": "BRL", "tipo": "kpi", "nivel": 2},
    {"linha": 19, "categoria": "        Ads", "unidade": "BRL", "tipo": "kpi", "nivel": 2},
    {"linha": 20, "categoria": "    USA", "unidade": "BRL", "tipo": "kpi", "nivel": 1},
    {"linha": 21, "categoria": "        EaM", "unidade": "BRL", "tipo": "kpi", "nivel": 2},
    {"linha": 22, "categoria": "        Ads", "unidade": "BRL", "tipo": "kpi", "nivel": 2},
    {"linha": 23, "categoria": "", "unidade": "", "tipo": "blank", "nivel": 0},
    
    # RECEITA
    {"linha": 24, "categoria": "Revenue", "unidade": "BRL", "tipo": "revenue", "nivel": 0, "formula": "sum"},
    {"linha": 25, "categoria": "Revenue no Tax (total das receitas recebidas em conta)", "unidade": "BRL", "tipo": "revenue", "nivel": 0, "formula": "sum"},
    {"linha": 26, "categoria": "     Google (IAPHUB)", "unidade": "BRL", "tipo": "revenue", "nivel": 1, "formula": "import"},
    {"linha": 27, "categoria": "     Google (IAPHUB) x 0.85", "unidade": "BRL", "tipo": "revenue", "nivel": 1, "formula": "calc"},
    {"linha": 28, "categoria": "          Δ(%)", "unidade": "%", "tipo": "kpi", "nivel": 2, "formula": "calc"},
    {"linha": 29, "categoria": "     Google", "unidade": "BRL", "tipo": "revenue", "nivel": 1, "formula": "import"},
    {"linha": 30, "categoria": "          Brazil", "unidade": "BRL", "tipo": "revenue", "nivel": 2, "formula": "import"},
    {"linha": 31, "categoria": "          Brazil Margin (%)", "unidade": "%", "tipo": "kpi", "nivel": 2, "formula": "calc"},
    {"linha": 32, "categoria": "          USA", "unidade": "BRL", "tipo": "revenue", "nivel": 2, "formula": "import"},
    {"linha": 33, "categoria": "          USA Margin (%)", "unidade": "%", "tipo": "kpi", "nivel": 2, "formula": "calc"},
    {"linha": 34, "categoria": "     Apple (IAPHUB)", "unidade": "BRL", "tipo": "revenue", "nivel": 1, "formula": "import"},
    {"linha": 35, "categoria": "     Apple (IAPHUB) x 0.85", "unidade": "BRL", "tipo": "revenue", "nivel": 1, "formula": "calc"},
    {"linha": 36, "categoria": "          Δ(%)", "unidade": "%", "tipo": "kpi", "nivel": 2, "formula": "calc"},
    {"linha": 37, "categoria": "     Apple", "unidade": "BRL", "tipo": "revenue", "nivel": 1, "formula": "import"},
    {"linha": 38, "categoria": "          Brazil", "unidade": "BRL", "tipo": "revenue", "nivel": 2, "formula": "import"},
    {"linha": 39, "categoria": "          Brazil Margin (%)", "unidade": "%", "tipo": "kpi", "nivel": 2, "formula": "calc"},
    {"linha": 40, "categoria": "          USA", "unidade": "BRL", "tipo": "revenue", "nivel": 2, "formula": "import"},
    {"linha": 41, "categoria": "          USA Margin (%)", "unidade": "%", "tipo": "kpi", "nivel": 2, "formula": "calc"},
    {"linha": 42, "categoria": "Invest Income (rendimento de aplicação & cashback)", "unidade": "BRL", "tipo": "revenue", "nivel": 0, "formula": "import"},
    {"linha": 43, "categoria": "", "unidade": "", "tipo": "blank", "nivel": 0},
    
    # COSTS OF REVENUE
    {"linha": 44, "categoria": "Costs of Revenue", "unidade": "BRL", "tipo": "cost", "nivel": 0, "formula": "sum"},
    {"linha": 45, "categoria": "Payment Processing Expenses (despesas de processamento de pagamento)", "unidade": "BRL", "tipo": "cost", "nivel": 0, "formula": "calc"},
    {"linha": 46, "categoria": "COGS (CMV)", "unidade": "BRL", "tipo": "cost", "nivel": 0, "formula": "sum"},
    {"linha": 47, "categoria": "     AWS", "unidade": "BRL", "tipo": "cost", "nivel": 1, "formula": "import"},
    {"linha": 48, "categoria": "     Cloudflare", "unidade": "BRL", "tipo": "cost", "nivel": 1, "formula": "import"},
    {"linha": 49, "categoria": "     Heroku", "unidade": "BRL", "tipo": "cost", "nivel": 1, "formula": "import"},
    {"linha": 50, "categoria": "     IAPHUB", "unidade": "BRL", "tipo": "cost", "nivel": 1, "formula": "import"},
    {"linha": 51, "categoria": "     MailGun", "unidade": "BRL", "tipo": "cost", "nivel": 1, "formula": "import"},
    {"linha": 52, "categoria": "     AWS SES", "unidade": "BRL", "tipo": "cost", "nivel": 1, "formula": "import"},
    {"linha": 53, "categoria": "", "unidade": "", "tipo": "blank", "nivel": 0},
    
    # GROSS PROFIT
    {"linha": 54, "categoria": "Gross Profit", "unidade": "BRL", "tipo": "profit", "nivel": 0, "formula": "calc"},
    {"linha": 55, "categoria": "     Gross Margin", "unidade": "%", "tipo": "kpi", "nivel": 1, "formula": "calc"},
    {"linha": 56, "categoria": "", "unidade": "", "tipo": "blank", "nivel": 0},
    
    # OPERATING EXPENSES
    {"linha": 57, "categoria": "Operating Expenses (Despesas Administrativas e Vendas Gerais)", "unidade": "BRL", "tipo": "cost", "nivel": 0, "formula": "sum"},
    {"linha": 58, "categoria": "R&D", "unidade": "BRL", "tipo": "cost", "nivel": 0, "formula": "import"},
    {"linha": 59, "categoria": "SG&A", "unidade": "BRL", "tipo": "cost", "nivel": 0, "formula": "sum"},
    {"linha": 60, "categoria": "     Marketing", "unidade": "BRL", "tipo": "cost", "nivel": 1, "formula": "import"},
    {"linha": 61, "categoria": "          Marketing/Revenue no Tax", "unidade": "%", "tipo": "kpi", "nivel": 2, "formula": "calc"},
    {"linha": 62, "categoria": "          EaM - BR", "unidade": "BRL", "tipo": "cost", "nivel": 2, "formula": "import"},
    {"linha": 63, "categoria": "          EaM - US", "unidade": "BRL", "tipo": "cost", "nivel": 2, "formula": "import"},
    {"linha": 64, "categoria": "          Affiliates & Events", "unidade": "BRL", "tipo": "cost", "nivel": 2, "formula": "import"},
    {"linha": 65, "categoria": "          Google Ads", "unidade": "BRL", "tipo": "cost", "nivel": 2, "formula": "import"},
    {"linha": 66, "categoria": "          Apple Search Ads (ASA)", "unidade": "BRL", "tipo": "cost", "nivel": 2, "formula": "import"},
    {"linha": 67, "categoria": "          Meta Ads", "unidade": "BRL", "tipo": "cost", "nivel": 2, "formula": "import"},
    {"linha": 68, "categoria": "     Wages", "unidade": "BRL", "tipo": "cost", "nivel": 1, "formula": "import"},
    {"linha": 69, "categoria": "     Tech Support & Services", "unidade": "BRL", "tipo": "cost", "nivel": 1, "formula": "sum"},
    {"linha": 70, "categoria": "          Conta Azul", "unidade": "BRL", "tipo": "cost", "nivel": 2, "formula": "import"},
    {"linha": 71, "categoria": "          G Suite", "unidade": "BRL", "tipo": "cost", "nivel": 2, "formula": "import"},
    {"linha": 72, "categoria": "          Adobe Creative Cloud", "unidade": "BRL", "tipo": "cost", "nivel": 2, "formula": "import"},
]

# Preencher estrutura
for item in estrutura_pl:
    linha = item['linha']
    
    # Categoria
    ws_pl[f'A{linha}'] = item['categoria']
    ws_pl[f'B{linha}'] = item['unidade']
    
    # Aplicar estilos baseado no nível e tipo
    if item['tipo'] == 'blank':
        continue
    elif item['nivel'] == 0 and item['tipo'] in ['revenue', 'cost', 'profit']:
        ws_pl[f'A{linha}'].font = category_font
        if item['tipo'] == 'revenue':
            ws_pl[f'A{linha}'].fill = revenue_fill
        elif item['tipo'] == 'cost':
            ws_pl[f'A{linha}'].fill = cost_fill
        else:
            ws_pl[f'A{linha}'].fill = category_fill
    elif item['nivel'] == 1:
        ws_pl[f'A{linha}'].font = Font(bold=False, size=10)
        ws_pl[f'A{linha}'].fill = subcategory_fill
    
    # Bordas
    ws_pl[f'A{linha}'].border = thin_border
    ws_pl[f'B{linha}'].border = thin_border
    
    # Fórmulas para cada mês
    for col_idx in range(3, 21):  # Colunas C a T (18 meses)
        col_letter = get_column_letter(col_idx)
        cell = ws_pl[f'{col_letter}{linha}']
        cell.border = thin_border
        
        # Aplicar formato numérico
        if item['unidade'] == 'BRL':
            cell.number_format = 'R$ #,##0.00'
        elif item['unidade'] == '%':
            cell.number_format = '0.00%'
        elif item['unidade'] == 'Users':
            cell.number_format = '#,##0'
        
        # Aplicar fórmulas (placeholder por enquanto)
        if item.get('formula') == 'import':
            # Fórmula SUMIFS para importar do Extrato_Importado
            # Será implementada na próxima etapa
            cell.value = 0
            cell.fill = editable_fill
        elif item.get('formula') == 'calc':
            # Fórmulas de cálculo
            cell.value = 0
        elif item.get('formula') == 'sum':
            # Fórmulas de soma
            cell.value = 0

# Ajustar larguras
ws_pl.column_dimensions['A'].width = 60
ws_pl.column_dimensions['B'].width = 10
for col_idx in range(3, 21):
    col_letter = get_column_letter(col_idx)
    ws_pl.column_dimensions[col_letter].width = 15

print(f"   ✓ {len(estrutura_pl)} linhas criadas no P&L")

# ============================================================================
# 5. SALVAR WORKBOOK
# ============================================================================
print("\n5. Salvando workbook...")

output_path = '/home/ubuntu/Business_Plan_Umatch_Automatizado_v2.xlsx'
wb.save(output_path)

print(f"\n✓ Workbook salvo em: {output_path}")
print("\n" + "=" * 80)
print("FASE 2 CONCLUÍDA - Estrutura P&L criada")
print("=" * 80)
print("\nPróxima etapa:")
print("  - Implementar fórmulas SUMIFS para importação automática")
print("  - Implementar fórmulas de cálculo (margens, percentuais, etc.)")
